import functions.analysis as anly
from openpyxl import load_workbook
import xlsxwriter
# /srv/2-lkeb-17-dl01/syousefi/TestCode/EsophagusProject/sythesize_code/Log/synth-forked_synthesizing_net_rotate-1/


from shutil import copyfile
import os, sys, inspect
import matplotlib.pyplot as plt
import numpy as np
import tensorflow as tf
from functions.loss.perceptual_loss.vgg.vgg_feature_maker import vgg_feature_maker

currentdir = os.path.dirname(os.path.abspath(inspect.getfile(inspect.currentframe())))
parentdir = os.path.dirname(currentdir)
sys.path.insert(0, parentdir)
from functions.image_reader.read_data import _read_data
from functions.image_reader.image_class import image_class
from functions.loss.loss_fun import _loss_func
from functions.network.perceptual_densenet import _densenet

del currentdir, parentdir
max_angio = 2.99771428108
max_perf = 17.0151833445

import SimpleITK as sitk

eps = 1E-5
plot_tag = 'ssim'
# densnet_unet_config = [1, 3, 3, 3, 1]
# ct_cube_size = 255
# db_size1 = np.int32(ct_cube_size - 2)
# db_size2 = np.int32(db_size1 / 2)
# db_size3 = np.int32(db_size2 / 2)
# crop_size1 = np.int32(((db_size3 - 2) * 2 + 1.0))
# crop_size2 = np.int32((crop_size1 - 2) * 2 + 1)


in_dim = np.int32(117)
in_size0 = np.int32(0)
in_size1 = np.int32(in_dim)
in_size2 = np.int32(in_size1)  # conv stack
in_size3 = np.int32((in_size2 - 2))  # level_design1
in_size4 = np.int32(in_size3 / 2 - 2)  # downsampleing1+level_design2
in_size5 = np.int32(in_size4 / 2 - 2)  # downsampleing2+level_design3
crop_size0 = np.int32(0)
crop_size1 = np.int32(2 * in_size5 + 1)
crop_size2 = np.int32(2 * crop_size1 + 1)
# in_size6 = int(in_size5 / 2 -2)  # downsampleing2+level_design3
# crop_size0 = (0)
# crop_size1 = np.int8(2 * in_size5 + 1)
# crop_size2 = np.int8(2 * crop_size1  + 1)
final_layer = crop_size2
label_patchs_size = final_layer
patch_window = in_dim  # 89
#
# in_size0 = (0)
# in_size1 = (input_dim)
# in_size2 = (in_size1)  # conv stack
# in_size3 = ((in_size2))  # level_design1
# in_size4 = int(in_size3 / 2)  # downsampleing1+level_design2
# in_size5 = int(in_size4 / 2 - 4)  # downsampleing2+level_design3
# crop_size0 = (0)
# crop_size1 = (2 * in_size5 + 1)
# crop_size2 = (2 * (crop_size1 - 4) + 1)

gt_cube_size = final_layer


# gap = ct_cube_size - gtv_cube_size


def test_all_nets():
    data = 2

    Server = 'DL'
    if Server == 'DL':
        parent_path = '/srv/2-lkeb-17-dl01/syousefi/TestCode/EsophagusProject/sythesize_code/ASL_LOG/Log_perceptual/regularization/perceptual-0/'
        data_path = '/srv/2-lkeb-17-dl01/syousefi/TestCode/EsophagusProject/Data-01/BrainWeb_permutation00_low/'
    else:
        parent_path = '/exports/lkeb-hpc/syousefi/Code/'
        data_path = '/srv/2-lkeb-17-dl01/syousefi/TestCode/EsophagusProject/Data-01/BrainWeb_permutation00_low/'

    img_name = ''
    label_name = ''

    _rd = _read_data(data=data, reverse=False,
                     img_name=img_name, label_name=label_name,
                     dataset_path=data_path)
    '''read path of the images for train, test, and validation'''
    train_data, validation_data, test_data = _rd.read_data_path()

    chckpnt_dir = parent_path + 'unet_checkpoints/'
    result_path = parent_path + 'results/'
    batch_no = 1
    batch_no_validation = batch_no
    # label_patchs_size = 87#39  # 63
    # patch_window = 103#53  # 77#89
    if test_vali == 1:
        test_set = validation_data
    else:
        test_set = test_data
    # ===================================================================================
    img_row1 = tf.placeholder(tf.float32,
                              shape=[batch_no, patch_window, patch_window, patch_window, 1],
                              name='img_row1')
    img_row2 = tf.placeholder(tf.float32,
                              shape=[batch_no, patch_window, patch_window, patch_window, 1],
                              name='img_row2')
    img_row3 = tf.placeholder(tf.float32,
                              shape=[batch_no, patch_window, patch_window, patch_window, 1],
                              name='img_row3')
    img_row4 = tf.placeholder(tf.float32,
                              shape=[batch_no, patch_window, patch_window, patch_window, 1],
                              name='img_row4')
    img_row5 = tf.placeholder(tf.float32,
                              shape=[batch_no, patch_window, patch_window, patch_window, 1],
                              name='img_row5')
    img_row6 = tf.placeholder(tf.float32,
                              shape=[batch_no, patch_window, patch_window, patch_window, 1],
                              name='img_row6')
    img_row7 = tf.placeholder(tf.float32,
                              shape=[batch_no, patch_window, patch_window, patch_window, 1],
                              name='img_row7')
    img_row8 = tf.placeholder(tf.float32,
                              shape=[batch_no, patch_window, patch_window, patch_window, 1],
                              name='img_row8')

    mri_ph = tf.placeholder(tf.float32,
                            shape=[batch_no, patch_window, patch_window, patch_window, 1],
                            name='mri')

    label1 = tf.placeholder(tf.float32, shape=[batch_no, label_patchs_size, label_patchs_size,
                                               label_patchs_size, 1], name='label1')
    label2 = tf.placeholder(tf.float32, shape=[batch_no, label_patchs_size, label_patchs_size,
                                               label_patchs_size, 1], name='label2')
    label3 = tf.placeholder(tf.float32, shape=[batch_no, label_patchs_size, label_patchs_size,
                                               label_patchs_size, 1], name='label3')
    label4 = tf.placeholder(tf.float32, shape=[batch_no, label_patchs_size, label_patchs_size,
                                               label_patchs_size, 1], name='label4')
    label5 = tf.placeholder(tf.float32, shape=[batch_no, label_patchs_size, label_patchs_size,
                                               label_patchs_size, 1], name='label5')
    label6 = tf.placeholder(tf.float32, shape=[batch_no, label_patchs_size, label_patchs_size,
                                               label_patchs_size, 1], name='label6')
    label7 = tf.placeholder(tf.float32, shape=[batch_no, label_patchs_size, label_patchs_size,
                                               label_patchs_size, 1], name='label7')
    label8 = tf.placeholder(tf.float32, shape=[batch_no, label_patchs_size, label_patchs_size,
                                               label_patchs_size, 1], name='label8')
    label9 = tf.placeholder(tf.float32, shape=[batch_no, label_patchs_size, label_patchs_size,
                                               label_patchs_size, 1], name='label9')
    label10 = tf.placeholder(tf.float32, shape=[batch_no, label_patchs_size, label_patchs_size,
                                                label_patchs_size, 1], name='label10')
    label11 = tf.placeholder(tf.float32, shape=[batch_no, label_patchs_size, label_patchs_size,
                                                label_patchs_size, 1], name='label11')
    label12 = tf.placeholder(tf.float32, shape=[batch_no, label_patchs_size, label_patchs_size,
                                                label_patchs_size, 1], name='label12')
    label13 = tf.placeholder(tf.float32, shape=[batch_no, label_patchs_size, label_patchs_size,
                                                label_patchs_size, 1], name='label13')
    label14 = tf.placeholder(tf.float32, shape=[batch_no, label_patchs_size, label_patchs_size,
                                                label_patchs_size, 1], name='label14')

    is_training = tf.placeholder(tf.bool, name='is_training')
    input_dim = tf.placeholder(tf.int32, name='input_dim')

    perf_vgg_loss_tens = tf.placeholder(tf.float32, name='VGG_perf')
    angio_vgg_loss_tens = tf.placeholder(tf.float32, name='VGG_angio')
    perf_vgg_tens0 = tf.placeholder(tf.float32, name='vgg_perf0')
    perf_vgg_tens1 = tf.placeholder(tf.float32, name='vgg_perf1')
    perf_vgg_tens2 = tf.placeholder(tf.float32, name='vgg_perf2')
    perf_vgg_tens3 = tf.placeholder(tf.float32, name='vgg_perf3')
    perf_vgg_tens4 = tf.placeholder(tf.float32, name='vgg_perf4')
    perf_vgg_tens5 = tf.placeholder(tf.float32, name='vgg_perf5')
    perf_vgg_tens6 = tf.placeholder(tf.float32, name='vgg_perf6')

    angio_vgg_tens0 = tf.placeholder(tf.float32, name='vgg_angio0')
    angio_vgg_tens1 = tf.placeholder(tf.float32, name='vgg_angio1')
    angio_vgg_tens2 = tf.placeholder(tf.float32, name='vgg_angio2')
    angio_vgg_tens3 = tf.placeholder(tf.float32, name='vgg_angio3')
    angio_vgg_tens4 = tf.placeholder(tf.float32, name='vgg_angio4')
    angio_vgg_tens5 = tf.placeholder(tf.float32, name='vgg_angio5')
    angio_vgg_tens6 = tf.placeholder(tf.float32, name='vgg_angio6')
    perf_huber_loss_tens = tf.placeholder(tf.float32, name='huber_perf')
    angio_huber_loss_tens = tf.placeholder(tf.float32, name='huber_angio')

    perf_huber_tens0 = tf.placeholder(tf.float32, name='huber_perf0')
    perf_huber_tens1 = tf.placeholder(tf.float32, name='huber_perf1')
    perf_huber_tens2 = tf.placeholder(tf.float32, name='huber_perf2')
    perf_huber_tens3 = tf.placeholder(tf.float32, name='huber_perf3')
    perf_huber_tens4 = tf.placeholder(tf.float32, name='huber_perf4')
    perf_huber_tens5 = tf.placeholder(tf.float32, name='huber_perf5')
    perf_huber_tens6 = tf.placeholder(tf.float32, name='huber_perf6')

    angio_huber_tens0 = tf.placeholder(tf.float32, name='huber_angio0')
    angio_huber_tens1 = tf.placeholder(tf.float32, name='huber_angio1')
    angio_huber_tens2 = tf.placeholder(tf.float32, name='huber_angio2')
    angio_huber_tens3 = tf.placeholder(tf.float32, name='huber_angio3')
    angio_huber_tens4 = tf.placeholder(tf.float32, name='huber_angio4')
    angio_huber_tens5 = tf.placeholder(tf.float32, name='huber_angio5')
    angio_huber_tens6 = tf.placeholder(tf.float32, name='huber_angio6')
    # ===================================================================================
    densenet = _densenet()

    [y, _] = densenet.densenet(img_row1=img_row1, img_row2=img_row2, img_row3=img_row3, img_row4=img_row4,
                               img_row5=img_row5,
                               img_row6=img_row6, img_row7=img_row7, img_row8=img_row8, input_dim=input_dim,
                               is_training=is_training)
    vgg = vgg_feature_maker(test=1)
    feature_type = 'huber'
    vgg_y0 = vgg.feed_img(y[:, :, :, :, 0], feature_type=feature_type).copy()
    vgg_y1 = vgg.feed_img(y[:, :, :, :, 1], feature_type=feature_type).copy()
    vgg_y2 = vgg.feed_img(y[:, :, :, :, 2], feature_type=feature_type).copy()
    vgg_y3 = vgg.feed_img(y[:, :, :, :, 3], feature_type=feature_type).copy()
    vgg_y4 = vgg.feed_img(y[:, :, :, :, 4], feature_type=feature_type).copy()
    vgg_y5 = vgg.feed_img(y[:, :, :, :, 5], feature_type=feature_type).copy()
    vgg_y6 = vgg.feed_img(y[:, :, :, :, 6], feature_type=feature_type).copy()

    vgg_y7 = vgg.feed_img(y[:, :, :, :, 7], feature_type=feature_type)
    vgg_y8 = vgg.feed_img(y[:, :, :, :, 8], feature_type=feature_type)
    vgg_y9 = vgg.feed_img(y[:, :, :, :, 9], feature_type=feature_type)
    vgg_y10 = vgg.feed_img(y[:, :, :, :, 10], feature_type=feature_type)
    vgg_y11 = vgg.feed_img(y[:, :, :, :, 11], feature_type=feature_type)
    vgg_y12 = vgg.feed_img(y[:, :, :, :, 12], feature_type=feature_type)
    vgg_y13 = vgg.feed_img(y[:, :, :, :, 13], feature_type=feature_type)

    vgg_label0 = vgg.feed_img(label1[:, :, :, :, 0], feature_type=feature_type).copy()
    vgg_label1 = vgg.feed_img(label2[:, :, :, :, 0], feature_type=feature_type).copy()
    vgg_label2 = vgg.feed_img(label3[:, :, :, :, 0], feature_type=feature_type).copy()
    vgg_label3 = vgg.feed_img(label4[:, :, :, :, 0], feature_type=feature_type).copy()
    vgg_label4 = vgg.feed_img(label5[:, :, :, :, 0], feature_type=feature_type).copy()
    vgg_label5 = vgg.feed_img(label6[:, :, :, :, 0], feature_type=feature_type).copy()
    vgg_label6 = vgg.feed_img(label7[:, :, :, :, 0], feature_type=feature_type).copy()

    vgg_label7 = vgg.feed_img(label8[:, :, :, :, 0], feature_type=feature_type)
    vgg_label8 = vgg.feed_img(label9[:, :, :, :, 0], feature_type=feature_type)
    vgg_label9 = vgg.feed_img(label10[:, :, :, :, 0], feature_type=feature_type)
    vgg_label10 = vgg.feed_img(label11[:, :, :, :, 0], feature_type=feature_type)
    vgg_label11 = vgg.feed_img(label12[:, :, :, :, 0], feature_type=feature_type)
    vgg_label12 = vgg.feed_img(label13[:, :, :, :, 0], feature_type=feature_type)
    vgg_label13 = vgg.feed_img(label14[:, :, :, :, 0], feature_type=feature_type)

    all_loss = tf.placeholder(tf.float32, name='loss')
    # is_training = tf.placeholder(tf.bool, name='is_training')
    # input_dim = tf.placeholder(tf.int32, name='input_dim')
    # ave_huber = tf.placeholder(tf.float32, name='huber')


    labels = []
    labels.append(label1)
    labels.append(label2)
    labels.append(label3)
    labels.append(label4)
    labels.append(label5)
    labels.append(label6)
    labels.append(label7)

    labels.append(label8)
    labels.append(label9)
    labels.append(label10)
    labels.append(label11)
    labels.append(label12)
    labels.append(label13)
    labels.append(label14)

    logits = []
    logits.append(y[:, :, :, :, 0, np.newaxis])
    logits.append(y[:, :, :, :, 1, np.newaxis])
    logits.append(y[:, :, :, :, 2, np.newaxis])
    logits.append(y[:, :, :, :, 3, np.newaxis])
    logits.append(y[:, :, :, :, 4, np.newaxis])
    logits.append(y[:, :, :, :, 5, np.newaxis])
    logits.append(y[:, :, :, :, 6, np.newaxis])

    logits.append(y[:, :, :, :, 7, np.newaxis])
    logits.append(y[:, :, :, :, 8, np.newaxis])
    logits.append(y[:, :, :, :, 9, np.newaxis])
    logits.append(y[:, :, :, :, 10, np.newaxis])
    logits.append(y[:, :, :, :, 11, np.newaxis])
    logits.append(y[:, :, :, :, 12, np.newaxis])
    logits.append(y[:, :, :, :, 13, np.newaxis])

    loss_instance = _loss_func()
    vgg_in_feature = []
    vgg_in_feature.append(vgg_y0)
    vgg_in_feature.append(vgg_y1)
    vgg_in_feature.append(vgg_y2)
    vgg_in_feature.append(vgg_y3)
    vgg_in_feature.append(vgg_y4)
    vgg_in_feature.append(vgg_y5)
    vgg_in_feature.append(vgg_y6)

    vgg_in_feature.append(vgg_y7)
    vgg_in_feature.append(vgg_y8)
    vgg_in_feature.append(vgg_y9)
    vgg_in_feature.append(vgg_y10)
    vgg_in_feature.append(vgg_y11)
    vgg_in_feature.append(vgg_y12)
    vgg_in_feature.append(vgg_y13)

    vgg_label_feature = []
    vgg_label_feature.append(vgg_label0)
    vgg_label_feature.append(vgg_label1)
    vgg_label_feature.append(vgg_label2)
    vgg_label_feature.append(vgg_label3)
    vgg_label_feature.append(vgg_label4)
    vgg_label_feature.append(vgg_label5)
    vgg_label_feature.append(vgg_label6)

    vgg_label_feature.append(vgg_label7)
    vgg_label_feature.append(vgg_label8)
    vgg_label_feature.append(vgg_label9)
    vgg_label_feature.append(vgg_label10)
    vgg_label_feature.append(vgg_label11)
    vgg_label_feature.append(vgg_label12)
    vgg_label_feature.append(vgg_label13)
    with tf.name_scope('Loss'):
        loss_dic = loss_instance.loss_selector('content_vgg_pairwise_loss_huber',
                                               labels=vgg_label_feature, logits=vgg_in_feature, vgg=vgg,
                                               h_labels=labels, h_logits=logits)
        cost = tf.reduce_mean(loss_dic["loss"], name="cost")
        # cost_angio = tf.reduce_mean(loss_dic["angio_SSIM"], name="angio_SSIM")
        # cost_perf = tf.reduce_mean(loss_dic["perf_SSIM"], name="perf_SSIM")

    # ========================================================================
    # ave_loss = tf.placeholder(tf.float32, name='loss')
    # ave_loss_perf = tf.placeholder(tf.float32, name='loss_perf')
    # ave_loss_angio = tf.placeholder(tf.float32, name='loss_angio')
    #
    # average_gradient_perf = tf.placeholder(tf.float32, name='grad_ave_perf')
    # average_gradient_angio = tf.placeholder(tf.float32, name='grad_ave_angio')
    #
    # ave_huber = tf.placeholder(tf.float32, name='huber')
    # restore the model
    sess = tf.Session()
    saver = tf.train.Saver()

    ckpt = tf.train.get_checkpoint_state(chckpnt_dir)
    saver.restore(sess, ckpt.model_checkpoint_path)

    copyfile('./test_synthesize_ssim_perf_angio.py', result_path + '/test_synthesize_ssim_perf_angio.py')

    _image_class = image_class(train_data
                               , bunch_of_images_no=1,
                               is_training=1,
                               patch_window=patch_window,
                               sample_no_per_bunch=1,
                               label_patch_size=label_patchs_size,
                               validation_total_sample=0)
    learning_rate = 1E-5
    extra_update_ops = tf.get_collection(tf.GraphKeys.UPDATE_OPS)
    with tf.control_dependencies(extra_update_ops):
        optimizer = tf.train.AdamOptimizer(learning_rate).minimize(cost)
        # init = tf.global_variables_initializer()
    dic_perf0 = []
    dic_perf1 = []
    dic_perf2 = []
    dic_perf3 = []
    dic_perf4 = []
    dic_perf5 = []
    dic_perf6 = []

    dic_angio0 = []
    dic_angio1 = []
    dic_angio2 = []
    dic_angio3 = []
    dic_angio4 = []
    dic_angio5 = []
    dic_angio6 = []
    loss = 0
    for img_indx in range(len(test_set)):
        crush, noncrush, perf, angio, mri, segmentation_, spacing, direction, origin = _image_class.read_image_for_test(
            test_set=test_set, img_indx=img_indx, input_size=in_dim, final_layer=final_layer)
        [out] = sess.run([y],
                      feed_dict={img_row1: np.expand_dims(np.expand_dims(crush[0], 0), -1),
                                img_row2: np.expand_dims(np.expand_dims(noncrush[1], 0), -1),
                                img_row3: np.expand_dims(np.expand_dims(crush[2], 0), -1),
                                img_row4: np.expand_dims(np.expand_dims(noncrush[3], 0), -1),
                                img_row5: np.expand_dims(np.expand_dims(crush[4], 0), -1),
                                img_row6: np.expand_dims(np.expand_dims(noncrush[5], 0), -1),
                                img_row7: np.expand_dims(np.expand_dims(crush[6], 0), -1),
                                img_row8: np.expand_dims(np.expand_dims(noncrush[7], 0), -1),
                                mri_ph: np.expand_dims(np.expand_dims(mri, 0), -1),
                                label1: np.expand_dims(np.expand_dims(perf[0], 0), -1),
                                label2: np.expand_dims(np.expand_dims(perf[1], 0), -1),
                                label3: np.expand_dims(np.expand_dims(perf[2], 0), -1),
                                label4: np.expand_dims(np.expand_dims(perf[3], 0), -1),
                                label5: np.expand_dims(np.expand_dims(perf[4], 0), -1),
                                label6: np.expand_dims(np.expand_dims(perf[5], 0), -1),
                                label7: np.expand_dims(np.expand_dims(perf[6], 0), -1),
                                label8: np.expand_dims(np.expand_dims(angio[0], 0), -1),
                                label9: np.expand_dims(np.expand_dims(angio[1], 0), -1),
                                label10: np.expand_dims(np.expand_dims(angio[2], 0), -1),
                                label11: np.expand_dims(np.expand_dims(angio[3], 0), -1),
                                label12: np.expand_dims(np.expand_dims(angio[4], 0), -1),
                                label13: np.expand_dims(np.expand_dims(angio[5], 0), -1),
                                label14: np.expand_dims(np.expand_dims(angio[6], 0), -1),
                                is_training: False,
                                input_dim: patch_window,
                                all_loss: -1.,

                                angio_vgg_loss_tens: -1,  # vgg angio
                                perf_vgg_loss_tens: -1,

                                perf_vgg_tens0: -1,
                                perf_vgg_tens1: -1,
                                perf_vgg_tens2: -1,
                                perf_vgg_tens3: -1,
                                perf_vgg_tens4: -1,
                                perf_vgg_tens5: -1,
                                perf_vgg_tens6: -1,
                                angio_vgg_tens0: -1,
                                angio_vgg_tens1: -1,
                                angio_vgg_tens2: -1,
                                angio_vgg_tens3: -1,
                                angio_vgg_tens4: -1,
                                angio_vgg_tens5: -1,
                                angio_vgg_tens6: -1,

                                perf_huber_loss_tens: -1,
                                angio_huber_loss_tens: -1,

                                perf_huber_tens0: -1,
                                perf_huber_tens1: -1,
                                perf_huber_tens2: -1,
                                perf_huber_tens3: -1,
                                perf_huber_tens4: -1,
                                perf_huber_tens5: -1,
                                perf_huber_tens6: -1,

                                angio_huber_tens0: -1,
                                angio_huber_tens1: -1,
                                angio_huber_tens2: -1,
                                angio_huber_tens3: -1,
                                angio_huber_tens4: -1,
                                angio_huber_tens5: -1,
                                angio_huber_tens6: -1,
                                })

        for i in range(np.shape(out)[-1]):
            image = out[0, :, :, :, i]
            sitk_image = sitk.GetImageFromArray(image)
            res_dir = test_set[img_indx][0][0].split('/')[-2]
            if i == 0:
                os.mkdir(parent_path + 'results/' + res_dir)
            if i < 7:
                nm = 'perf'
            else:
                nm = 'angi'
            sitk_image.SetDirection(direction=direction)
            sitk_image.SetOrigin(origin=origin)
            sitk_image.SetSpacing(spacing=spacing)
            sitk.WriteImage(sitk_image, parent_path + 'results/' + res_dir + '/' + nm + '_' + str(i % 7) + '.mha')
            print(parent_path + 'results/' + res_dir + '/' + nm + '_' + str(i % 7) + '.mha done!')
        for i in range(7):
            if i == 0:
                os.mkdir(parent_path + 'results/' + res_dir + '/GT/')
            sitk_angio = sitk.GetImageFromArray(angio[i])
            sitk_angio.SetDirection(direction=direction)
            sitk_angio.SetOrigin(origin=origin)
            sitk_angio.SetSpacing(spacing=spacing)
            sitk.WriteImage(sitk_angio, parent_path + 'results/' + res_dir + '/GT/angio_' + str(i) + '.mha')

            sitk_perf = sitk.GetImageFromArray(perf[i])
            sitk_perf.SetDirection(direction=direction)
            sitk_perf.SetOrigin(origin=origin)
            sitk_perf.SetSpacing(spacing=spacing)
            sitk.WriteImage(sitk_perf, parent_path + 'results/' + res_dir + '/GT/perf_' + str(i) + '.mha')
        a = 1
        dic_perf0.append(anly.analysis(out[0, :, :, :, 0], perf[i], 0, max_perf))
        dic_perf1.append(anly.analysis(out[0, :, :, :, 1], perf[i], 0, max_perf))
        dic_perf2.append(anly.analysis(out[0, :, :, :, 2], perf[i], 0, max_perf))
        dic_perf3.append(anly.analysis(out[0, :, :, :, 3], perf[i], 0, max_perf))
        dic_perf4.append(anly.analysis(out[0, :, :, :, 4], perf[i], 0, max_perf))
        dic_perf5.append(anly.analysis(out[0, :, :, :, 5], perf[i], 0, max_perf))
        dic_perf6.append(anly.analysis(out[0, :, :, :, 6], perf[i], 0, max_perf))

        dic_angio0.append(anly.analysis(out[0, :, :, :, 7], angio[i], 0, max_angio))
        dic_angio1.append(anly.analysis(out[0, :, :, :, 8], angio[i], 0, max_angio))
        dic_angio2.append(anly.analysis(out[0, :, :, :, 9], angio[i], 0, max_angio))
        dic_angio3.append(anly.analysis(out[0, :, :, :, 10], angio[i], 0, max_angio))
        dic_angio4.append(anly.analysis(out[0, :, :, :, 11], angio[i], 0, max_angio))
        dic_angio5.append(anly.analysis(out[0, :, :, :, 12], angio[i], 0, max_angio))
        dic_angio6.append(anly.analysis(out[0, :, :, :, 13], angio[i], 0, max_angio))
        if img_indx == 0:
            headers = dic_perf0[0].keys()
        dics = [dic_perf0, dic_perf1, dic_perf2, dic_perf3, dic_perf4, dic_perf5, dic_perf6,
                dic_angio0, dic_angio1, dic_angio2, dic_angio3, dic_angio4, dic_angio5, dic_angio6]
    save_in_xlsx(parent_path, headers,
                 dics=dics)
        # plt.imshow(out[0, int(gt_cube_size / 2), :, :, 0])
        # plt.figure()
        # loss += loss_train1
        # print('Loss_train: ', loss_train1)

    print('Total loss: ', loss / len(test_set))
def save_in_xlsx(parent_path,headers,dics):
    # create excel file if it does not exist
    sheets=['perf0','perf1','perf2','perf3','perf4','perf5','perf6',
            'angio0','angio1','angio2','angio3','angio4','angio5','angio6']
    xlsx_name=parent_path+'results/results.xlsx'
    if not os.path.isfile(xlsx_name):
        book = xlsxwriter.Workbook(xlsx_name)
        for sh in sheets:
            sheet = book.add_worksheet(sh)
            for (idx, header) in enumerate(headers):
                sheet.write(0, idx, header)
        book.close()
    # loop through all dictionaries

    i=0
    book = load_workbook(xlsx_name)
    for dd in dics:
        sheet = book.get_sheet_by_name(sheets[i])
        i+=1
        for d in dd:
            values = [d[key] for key in headers]
            # write to excel file
            sheet.append(values)
    book.save(filename=xlsx_name)

if __name__ == "__main__":
    Log = 'EsophagusProject/sythesize_code/ASL_LOG/Log_perceptual/regularization/'
    fold = 0
    log_tag = 'perceptual-' + str(fold) + '/'
    test_vali = 0
    if test_vali == 1:
        out_dir = log_tag + '/result_vali/'
    else:
        out_dir = log_tag + '/results/'
    test_all_nets()