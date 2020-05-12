import functions.analysis as anly
from openpyxl import load_workbook
import xlsxwriter
import pandas
# /srv/2-lkeb-17-dl01/syousefi/TestCode/EsophagusProject/sythesize_code/Log/synth-forked_synthesizing_net_rotate-1/
from os import listdir
from os.path import isfile, join
from random import shuffle

from shutil import copyfile
import os, sys, inspect
import matplotlib.pyplot as plt
import numpy as np
import tensorflow as tf
import SimpleITK as sitk
currentdir = os.path.dirname(os.path.abspath(inspect.getfile(inspect.currentframe())))
parentdir = os.path.dirname(currentdir)
sys.path.insert(0, parentdir)
from functions.image_reader.read_data import _read_data
from functions.image_reader.image_class import image_class
from functions.loss.loss_fun import _loss_func
from functions.network.densenet_angio_perf import _densenet

def save_in_xlsx(parent_path, headers, dics,resultsdir):
    # create excel file if it does not exist
    sheets = ['perf0', 'perf1', 'perf2', 'perf3', 'perf4', 'perf5', 'perf6',
              'angio0', 'angio1', 'angio2', 'angio3', 'angio4', 'angio5', 'angio6']
    xlsx_name = parent_path + resultsdir+ 'results.xlsx'
    if not os.path.isfile(xlsx_name):
        book = xlsxwriter.Workbook(xlsx_name)
        for sh in sheets:
            sheet = book.add_worksheet(sh)
            for (idx, header) in enumerate(headers):
                sheet.write(0, idx, header)
        book.close()
    # loop through all dictionaries

    i = 0
    book = load_workbook(xlsx_name)
    for dd in dics:
        sheet = book.get_sheet_by_name(sheets[i])
        i += 1
        for d in dd:
            values = [d[key] for key in headers]
            # write to excel file
            sheet.append(values)
    book.save(filename=xlsx_name)
if __name__ == "__main__":
    resultsdir="results/"
    # image_path = "/exports/lkeb-hpc/syousefi/2-lkeb-17-dl01/syousefi/TestCode/EsophagusProject/sythesize_code/ASL_LOG/Workshop_log/MSE/experiment-2/"+resultsdir
    # image_path = "/exports/lkeb-hpc/syousefi/2-lkeb-17-dl01/syousefi/TestCode/EsophagusProject/sythesize_code/ASL_LOG/normal_SSIM_p_a_newdb/experiment-1/"+resultsdir
    # image_path = "/exports/lkeb-hpc/syousefi/2-lkeb-17-dl01/syousefi/TestCode/EsophagusProject/sythesize_code/ASL_LOG/multi_stage/experiment-2/"+resultsdir
    image_path = "/exports/lkeb-hpc/syousefi/2-lkeb-17-dl01/syousefi/TestCode/EsophagusProject/sythesize_code/ASL_LOG/Log_perceptual/regularization/perceptual-0/"+resultsdir
    data_dir = [join(image_path, f) for f in listdir(image_path) if ~isfile(join(image_path, f)) and not f.startswith(".") and not f.endswith(".py")and not f.endswith(".xlsx") ]
    dic_perf0 = []
    dic_perf1 = []
    dic_perf2 = []
    dic_perf3 = []
    dic_perf4 = []
    dic_perf5 = []
    dic_perf6 = []

    dic_angio0 = []
    dic_angio1 = []
    dic_angio2 = []
    dic_angio3 = []
    dic_angio4 = []
    dic_angio5 = []
    dic_angio6 = []
    max_angio = 2.99771428108
    max_perf = 17.0151833445
    type=["/angi_","/perf_"]
    gt_type=["/angio_","/perf_"]
    for i in range(2,len(data_dir)):
        for j in range(7):
            for k in range(2):
                output= sitk.GetArrayFromImage(sitk.ReadImage(data_dir[i] + type[k] +str(j)+ ".mha" ))
                GT= sitk.GetArrayFromImage(sitk.ReadImage(data_dir[i] +"/GT"+ gt_type[k] +str(j)+ ".mha" ))
                if k==1:
                    if j==0:
                        dic_perf0.append(anly.analysis(output,GT, 0, max_perf))
                    elif j==1:
                        dic_perf1.append(anly.analysis(output, GT, 0, max_perf))
                    elif j == 2:
                        dic_perf2.append(anly.analysis(output, GT, 0, max_perf))
                    elif j==3:
                        dic_perf3.append(anly.analysis(output, GT, 0, max_perf))
                    elif j==4:
                        dic_perf4.append(anly.analysis(output, GT, 0, max_perf))
                    elif j==5:
                        dic_perf5.append(anly.analysis(output, GT, 0, max_perf))
                    elif j==6:
                        dic_perf6.append(anly.analysis(output, GT, 0, max_perf))
                else:
                    if j==0:
                        dic_angio0.append(anly.analysis(output,GT, 0, max_angio))
                    elif j==1:
                        dic_angio1.append(anly.analysis(output, GT, 0, max_angio))
                    elif j == 2:
                        dic_angio2.append(anly.analysis(output, GT, 0, max_angio))
                    elif j==3:
                        dic_angio3.append(anly.analysis(output, GT, 0, max_angio))
                    elif j==4:
                        dic_angio4.append(anly.analysis(output, GT, 0, max_angio))
                    elif j==5:
                        dic_angio5.append(anly.analysis(output, GT, 0, max_angio))
                    elif j==6:
                        dic_angio6.append(anly.analysis(output, GT, 0, max_angio))
        if i == 2:
            headers = dic_perf0[0].keys()
        print(i)
    dics = [dic_perf0, dic_perf1, dic_perf2, dic_perf3, dic_perf4, dic_perf5, dic_perf6,
            dic_angio0, dic_angio1, dic_angio2, dic_angio3, dic_angio4, dic_angio5, dic_angio6]
    save_in_xlsx(image_path, headers,
                 dics=dics, resultsdir="")
